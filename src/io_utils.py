from __future__ import annotations

import io
from typing import BinaryIO, Union

import pandas as pd
import openpyxl


def load_absences_from_xlsx(file: Union[BinaryIO, io.BytesIO]) -> pd.DataFrame:
    """
    Reads the provided workbook and extracts absences from sheet 'Counter Sheet'
    columns:
      A: Date you left the uk
      B: Date yuo have returned
    Stops when column A is empty (after header).
    """
    # Streamlit UploadedFile behaves like a file-like object
    content = file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    if "Counter Sheet" not in wb.sheetnames:
        raise ValueError("Sheet 'Counter Sheet' not found in the uploaded workbook.")
    ws = wb["Counter Sheet"]

    leave_col = 1
    return_col = 2

    rows = []
    for r in range(2, ws.max_row + 1):
        leave = ws.cell(r, leave_col).value
        ret = ws.cell(r, return_col).value
        if leave is None and ret is None:
            # allow trailing blanks
            continue
        if leave is None:
            # skip malformed row
            continue
        rows.append({"leave_date": leave, "return_date": ret})

    return pd.DataFrame(rows)


def load_absences_from_csv(file: Union[BinaryIO, io.BytesIO]) -> pd.DataFrame:
    content = file.read()
    df = pd.read_csv(io.BytesIO(content))
    return df


def to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8")
